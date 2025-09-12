"""
STP Parser Module

Handles Excel file parsing and date format conversions for STP transactions.
"""

import logging
import io
import openpyxl
from datetime import datetime
from typing import List, Dict, Any, Optional, Union

logger = logging.getLogger(__name__)


def parse_excel_file(file_content: bytes, filename: str) -> List[Dict[str, Any]]:
    """Extract transactions from Excel file using flexible column mapping"""
    try:
        # Load Excel file
        workbook = openpyxl.load_workbook(io.BytesIO(file_content))
        worksheet = workbook.active
        
        # Find header row containing "Fecha Operación"
        header_row = None
        column_mapping = {}
        
        # Search for header row (typically around row 32)
        for row_num in range(25, min(50, worksheet.max_row + 1)):
            row_values = [cell.value for cell in worksheet[row_num]]
            
            # Check if this row contains our expected headers
            if any(cell and 'fecha operación' in str(cell).lower() for cell in row_values):
                header_row = row_num
                
                # Map column positions by header names
                expected_headers = [
                    'fecha operación', 'fecha liquidación', 'tipo operación', 
                    'concepto', 'clave de rastreo', 'cargos', 'abonos', 'saldos'
                ]
                
                for col_idx, cell_value in enumerate(row_values):
                    if cell_value:
                        cell_text = str(cell_value).lower().strip()
                        for expected_header in expected_headers:
                            if expected_header in cell_text:
                                column_mapping[expected_header] = col_idx
                                break
                break
        
        if not header_row:
            raise Exception("Header row with 'Fecha Operación' not found")
        
        # Verify all required columns are present
        required_columns = [
            'fecha operación', 'fecha liquidación', 'tipo operación', 
            'concepto', 'clave de rastreo', 'cargos', 'abonos', 'saldos'
        ]
        
        missing_columns = [col for col in required_columns if col not in column_mapping]
        if missing_columns:
            raise Exception(f"Missing required columns: {missing_columns}")
        
        # Extract transaction data
        transactions = []
        
        for row_num in range(header_row + 1, worksheet.max_row + 1):
            row_values = [cell.value for cell in worksheet[row_num]]
            
            # Skip empty rows
            if not any(row_values):
                continue
            
            # Extract data using column mapping
            try:
                transaction = {
                    'fecha_operacion': format_date_value(row_values[column_mapping['fecha operación']]),
                    'fecha_liquidacion': format_date_value(row_values[column_mapping['fecha liquidación']]),
                    'tipo_operacion': str(row_values[column_mapping['tipo operación']] or ''),
                    'concepto': str(row_values[column_mapping['concepto']] or ''),
                    'clave_rastreo': str(row_values[column_mapping['clave de rastreo']] or ''),
                    'cargos': format_numeric_value(row_values[column_mapping['cargos']]),
                    'abonos': format_numeric_value(row_values[column_mapping['abonos']]),
                    'saldos': format_numeric_value(row_values[column_mapping['saldos']]),
                    'file': filename
                }
                
                # Only add if at least one significant field has data
                if (transaction['fecha_operacion'] or transaction['tipo_operacion'] or 
                    transaction['concepto'] or transaction['clave_rastreo'] or
                    transaction['cargos'] or transaction['abonos']):
                    transactions.append(transaction)
                    
            except Exception as e:
                logger.warning(f"Error processing row {row_num} in {filename}: {e}")
                continue
        
        logger.info(f"Extracted {len(transactions)} transactions from {filename}")
        return transactions
        
    except Exception as e:
        logger.error(f"Error parsing Excel file {filename}: {e}")
        raise


def format_date_value(value: Any) -> Optional[str]:
    """Format date value to DD/MM/YYYY string"""
    if value is None:
        return None
    
    if isinstance(value, datetime):
        return value.strftime('%d/%m/%Y')
    elif isinstance(value, str):
        return value.strip() if value.strip() else None
    else:
        return str(value) if value else None


def format_numeric_value(value: Any) -> Optional[float]:
    """Format numeric value"""
    if value is None or value == '':
        return None
    
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def convert_date_format(fecha_operacion: str) -> Optional[str]:
    """Convert DD/MM/YYYY to YYYY-MM format"""
    try:
        if not fecha_operacion:
            return None
        
        # Handle DD/MM/YYYY format
        if '/' in fecha_operacion:
            parts = fecha_operacion.split('/')
            if len(parts) == 3:
                day, month, year = parts
                return f"{year}-{month.zfill(2)}"
        
        # Handle YYYY-MM-DD format (convert to YYYY-MM)
        if '-' in fecha_operacion and len(fecha_operacion.split('-')) == 3:
            return '-'.join(fecha_operacion.split('-')[:2])
            
        return None
    except Exception:
        return None


def convert_dd_mm_yyyy_to_yyyy_mm_dd(date_str: str) -> str:
    """Convert DD/MM/YYYY to YYYY-MM-DD for comparison"""
    try:
        if not date_str or '/' not in date_str:
            return date_str
        
        parts = date_str.split('/')
        if len(parts) == 3:
            day, month, year = parts
            return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
        
        return date_str
    except Exception:
        return date_str


def check_file_parsing_status(account_files: List[Dict[str, Any]], 
                            tracking_data: Dict[str, Any], 
                            account_number: str) -> List[Dict[str, Any]]:
    """Identify files that need parsing with better logging"""
    files_to_parse = []
    account_tracking = tracking_data.get(account_number, {})
    
    logger.info(f"Checking parsing status for {len(account_files)} files")
    logger.info(f"Account tracking has {len(account_tracking)} tracked files: {list(account_tracking.keys())}")
    
    for file_info in account_files:
        filename = file_info.get('filename')  # Use 'filename' key (not 'name')
        if not filename:
            logger.warning(f"File info missing filename: {file_info}")
            continue
            
        if not filename.endswith('.xlsx'):
            logger.info(f"Skipping non-Excel file: {filename}")
            continue
        
        file_last_modified = file_info.get('last_modified_formatted')
        tracked_info = account_tracking.get(filename, {})
        tracked_last_parsed = tracked_info.get('last_parsed')
        
        logger.info(f"Checking file: {filename}")
        logger.info(f"  File last modified: {file_last_modified}")
        logger.info(f"  Last parsed: {tracked_last_parsed}")
        
        # Parse if file is new or has been modified since last parse
        needs_parsing = False
        reason = ""
        
        if not tracked_last_parsed:
            needs_parsing = True
            reason = "never parsed"
        elif file_last_modified and tracked_last_parsed:
            # Convert both to comparable format
            try:
                # Parse the dates properly
                if 'T' in file_last_modified:
                    file_dt = datetime.fromisoformat(file_last_modified.replace('Z', '+00:00'))
                else:
                    file_dt = datetime.fromisoformat(file_last_modified)
                
                if 'T' in tracked_last_parsed:
                    tracked_dt = datetime.fromisoformat(tracked_last_parsed.replace('Z', '+00:00'))
                else:
                    tracked_dt = datetime.fromisoformat(tracked_last_parsed)
                
                if file_dt > tracked_dt:
                    needs_parsing = True
                    reason = f"modified since last parse ({file_dt} > {tracked_dt})"
                else:
                    reason = f"up to date ({file_dt} <= {tracked_dt})"
            except Exception as date_error:
                logger.warning(f"Date parsing error for {filename}: {date_error}")
                # If we can't parse dates, err on the side of parsing
                needs_parsing = True
                reason = "date parsing error - will parse to be safe"
        elif not file_last_modified:
            # If we don't have file modification date, parse anyway
            needs_parsing = True
            reason = "no modification date available"
        
        logger.info(f"  Decision: {'PARSE' if needs_parsing else 'SKIP'} - {reason}")
        
        if needs_parsing:
            files_to_parse.append(file_info)
    
    logger.info(f"Final result: {len(files_to_parse)} files need parsing out of {len(account_files)} total")
    return files_to_parse