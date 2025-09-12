"""
STP Analytics Module

Handles data analysis, record counting, filtering, and Excel export functionality.
"""

import logging
import io
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional
from modules.stp.stp_database import get_json_database

logger = logging.getLogger(__name__)


def convert_date_format(date_str: str) -> Optional[str]:
    """Convert DD/MM/YYYY to YYYY-MM format for grouping"""
    if not date_str or '/' not in date_str:
        return None
    
    try:
        parts = date_str.split('/')
        if len(parts) == 3:
            day, month, year = parts
            return f"{year}-{month.zfill(2)}"
    except Exception:
        pass
    
    return None


def get_monthly_record_counts(access_token: str, year: int) -> Dict[str, Dict[str, int]]:
    """Get record counts for each account and month based on filename analysis"""
    try:
        record_counts = {}
        accounts = ['646180559700000009', '646990403000000003', '646180403000000004']
        
        for account_number in accounts:
            record_counts[account_number] = {}
            
            # Initialize all months to 0
            for month in range(1, 13):
                month_key = f"{year}-{month:02d}"
                record_counts[account_number][month_key] = 0
            
            # Load actual database
            database = get_json_database(account_number, access_token)
            transactions = database.get('transactions', [])
            
            # Count transactions by filename/month
            for transaction in transactions:
                filename = transaction.get('file')
                if filename:
                    # Extract YYYYMM from filename: "ec-646180559700000009-202507.xlsx" -> "202507"
                    try:
                        # Split by '-' and get the third part (YYYYMM), then remove extension
                        filename_parts = filename.split('-')
                        if len(filename_parts) >= 3:
                            yyyymm_part = filename_parts[2].split('.')[0]  # Remove .xlsx extension
                            
                            # Convert YYYYMM to YYYY-MM format
                            if len(yyyymm_part) == 6 and yyyymm_part.isdigit():
                                file_year = yyyymm_part[:4]
                                file_month = yyyymm_part[4:6]
                                month_key = f"{file_year}-{file_month}"
                                
                                # Only count if it's for the requested year
                                if file_year == str(year):
                                    if month_key in record_counts[account_number]:
                                        record_counts[account_number][month_key] += 1
                                    else:
                                        # Add month if it doesn't exist (shouldn't happen with initialization)
                                        record_counts[account_number][month_key] = 1
                    except (IndexError, ValueError) as e:
                        logger.warning(f"Could not parse filename {filename}: {e}")
                        continue
        
        logger.info(f"Generated record counts for {len(accounts)} accounts in {year}: {record_counts}")
        return record_counts
        
    except Exception as e:
        logger.error(f"Error getting monthly record counts: {e}")
        return {}


def apply_export_filters(transactions: List[Dict[str, Any]], 
                        start_date: Optional[str], 
                        end_date: Optional[str], 
                        export_type: str) -> List[Dict[str, Any]]:
    """Apply date filters to transactions"""
    filtered = transactions.copy()
    
    if export_type == 'current_year':
        current_year = datetime.now().year
        filtered = [t for t in filtered if t.get('fecha_operacion_converted', '').startswith(str(current_year))]
    
    elif export_type == 'last_12_months':
        twelve_months_ago = datetime.now() - timedelta(days=365)
        cutoff_date = twelve_months_ago.strftime('%Y-%m-%d')
        filtered = [t for t in filtered if t.get('fecha_operacion_converted', '') >= cutoff_date]
    
    elif start_date and end_date:
        filtered = [t for t in filtered 
                   if start_date <= t.get('fecha_operacion_converted', '') <= end_date]
    
    elif start_date:
        filtered = [t for t in filtered if t.get('fecha_operacion_converted', '') >= start_date]
    
    elif end_date:
        filtered = [t for t in filtered if t.get('fecha_operacion_converted', '') <= end_date]
    
    return filtered


def create_formatted_excel(transactions: List[Dict[str, Any]], 
                          metadata: Dict[str, Any], 
                          export_type: str) -> bytes:
    """Create formatted Excel file from transactions"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions"
        
        # Headers
        headers = [
            "Fecha Operación", "Fecha Liquidación", "Tipo Operación", 
            "Concepto", "Clave de Rastreo", "Cargos", "Abonos", "Saldos", "File"
        ]
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for row, transaction in enumerate(transactions, 2):
            # Use original date format for display
            ws.cell(row=row, column=1, value=transaction.get('fecha_operacion_original', transaction.get('fecha_operacion')))
            ws.cell(row=row, column=2, value=transaction.get('fecha_liquidacion'))
            ws.cell(row=row, column=3, value=transaction.get('tipo_operacion'))
            ws.cell(row=row, column=4, value=transaction.get('concepto'))
            ws.cell(row=row, column=5, value=transaction.get('clave_rastreo'))
            ws.cell(row=row, column=6, value=transaction.get('cargos'))
            ws.cell(row=row, column=7, value=transaction.get('abonos'))
            ws.cell(row=row, column=8, value=transaction.get('saldos'))
            ws.cell(row=row, column=9, value=transaction.get('file'))
        
        # Format columns
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            
            # Auto-adjust column width
            ws.column_dimensions[column_letter].width = 20
            
            # Format date columns
            if col in [1, 2]:  # Date columns
                for row in range(2, len(transactions) + 2):
                    ws[f"{column_letter}{row}"].number_format = 'DD/MMM/YY'
            
            # Format number columns
            elif col in [6, 7, 8]:  # Cargos, Abonos, Saldos
                for row in range(2, len(transactions) + 2):
                    ws[f"{column_letter}{row}"].number_format = '#,##0.00'
        
        # Add autofilter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer.getvalue()
        
    except ImportError:
        logger.error("openpyxl library not available for Excel export")
        raise Exception("Excel export functionality not available")
    except Exception as e:
        logger.error(f"Error creating Excel file: {e}")
        raise


def calculate_transaction_statistics(transactions: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Calculate basic statistics for transactions"""
    try:
        if not transactions:
            return {
                'total_transactions': 0,
                'total_cargos': 0,
                'total_abonos': 0,
                'unique_files': 0,
                'date_range': None
            }
        
        total_cargos = sum(t.get('cargos', 0) or 0 for t in transactions)
        total_abonos = sum(t.get('abonos', 0) or 0 for t in transactions)
        unique_files = len(set(t.get('file') for t in transactions if t.get('file')))
        
        # Get date range
        dates = [t.get('fecha_operacion') for t in transactions if t.get('fecha_operacion')]
        date_range = None
        if dates:
            min_date = min(dates)
            max_date = max(dates)
            date_range = f"{min_date} to {max_date}" if min_date != max_date else min_date
        
        return {
            'total_transactions': len(transactions),
            'total_cargos': total_cargos,
            'total_abonos': total_abonos,
            'unique_files': unique_files,
            'date_range': date_range
        }
        
    except Exception as e:
        logger.error(f"Error calculating transaction statistics: {e}")
        return {
            'total_transactions': len(transactions) if transactions else 0,
            'total_cargos': 0,
            'total_abonos': 0,
            'unique_files': 0,
            'date_range': None
        }


def group_transactions_by_month(transactions: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    """Group transactions by month using filename analysis"""
    try:
        grouped = {}
        
        for transaction in transactions:
            filename = transaction.get('file')
            if filename:
                # Extract YYYYMM from filename and convert to YYYY-MM format
                try:
                    filename_parts = filename.split('-')
                    if len(filename_parts) >= 3:
                        yyyymm_part = filename_parts[2].split('.')[0]  # Remove extension
                        
                        if len(yyyymm_part) == 6 and yyyymm_part.isdigit():
                            file_year = yyyymm_part[:4]
                            file_month = yyyymm_part[4:6]
                            month_key = f"{file_year}-{file_month}"
                            
                            if month_key not in grouped:
                                grouped[month_key] = []
                            grouped[month_key].append(transaction)
                except (IndexError, ValueError):
                    # Skip transactions with invalid filenames
                    continue
        
        return grouped
        
    except Exception as e:
        logger.error(f"Error grouping transactions by month: {e}")
        return {}


def get_transaction_summary_by_type(transactions: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Get summary of transactions by operation type"""
    try:
        summary = {}
        
        for transaction in transactions:
            tipo_operacion = transaction.get('tipo_operacion', 'Unknown')
            
            if tipo_operacion not in summary:
                summary[tipo_operacion] = {
                    'count': 0,
                    'total_cargos': 0,
                    'total_abonos': 0
                }
            
            summary[tipo_operacion]['count'] += 1
            summary[tipo_operacion]['total_cargos'] += transaction.get('cargos', 0) or 0
            summary[tipo_operacion]['total_abonos'] += transaction.get('abonos', 0) or 0
        
        return summary
        
    except Exception as e:
        logger.error(f"Error creating transaction summary by type: {e}")
        return {}